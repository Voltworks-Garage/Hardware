import pygetwindow as gw
import pyautogui
import keyboard
import easygui
from threading import Thread
from time import sleep
import openpyxl


class refDesSearch:
    window_flag = True
    key_flag = False
    quit = False
    bom_headers = []
    bom_list = []
    component_index = 0
    window_location = None
    ok_to_relocate = False

    def __init__(self):
        # get BOM file by user input
        # f = easygui.fileopenbox(default="C:\Repos\E_Moto\Hardware")
        f = easygui.fileopenbox(default="C:\REPOS\E_Moto\Hardware\BMS\BMS_02\/")
        # load the excel workbook
        wb = openpyxl.load_workbook(filename=f)
        ws = wb.active
        # get the header row information
        for item in ws[1]:
            self.bom_headers.append(item.value)
        # create a dictionary of all the BOM items
        for row in range(2, ws.max_row + 1):
            this_dict = {}
            for num, param in enumerate(ws[row]):
                this_dict[self.bom_headers[num]] = str(param.value)
            self.bom_list.append(this_dict)

        self.bom_list = sorted(self.bom_list, key=lambda i: (i['MFG_PN'], i['RefDes']))

        ref_des_types = []
        for item in self.bom_list:
            type = self.get_ref_des_type(item["RefDes"])
            if type not in ref_des_types:
                ref_des_types.append(type)

        new_bom_list = []
        for red_des_type in sorted(ref_des_types):
            for item in self.bom_list:
                if item["RefDes"].startswith(red_des_type):
                    new_bom_list.append(item)

        self.bom_list = new_bom_list

    @staticmethod
    def get_ref_des_type(ref_des):
        ret_char = ''
        for char in ref_des:
            if char.isalpha():
                ret_char += char
            else:
                break
        return ret_char

    def show_window(self):
        while True:
            sleep(.1)
            if self.quit:
                return
            if self.window_flag is True:
                string = ""
                for item in self.bom_list[self.component_index]:
                    string += item + ": " + self.bom_list[self.component_index][item] + "\n"
                string += "\n\nUse < and > keys to cycle through components. Press Cancel or q to quit"
                if easygui.ccbox(string, 'BOM_ITEM') is False:
                    self.quit = True
                    return
                self.window_flag = False
                self.ok_to_relocate = False

    def relocate_window(self):
        while True:
            sleep(.1)
            win = gw.getWindowsWithTitle('BOM_ITEM')
            if len(win):
                if self.ok_to_relocate:
                    win[0].activate()
                    current_location = [win[0].left, win[0].top]
                    if self.window_location is not None and current_location[0] != self.window_location[0]:
                        win[0].moveTo(self.window_location[0], self.window_location[1])
                        self.ok_to_relocate = False

    def close_window(self):
        win = gw.getWindowsWithTitle('BOM_ITEM')
        if len(win):
            self.window_location = [win[0].left, win[0].top]
            try:
                win[0].activate()
            except:
                pass
            pyautogui.press("esc")
            self.ok_to_relocate = True

    def keyboard_input(self):
        self.key_flag = False
        while True:
            sleep(.001)
            if gw.getActiveWindowTitle() == 'BOM_ITEM':
                if self.quit:
                    return
                if keyboard.is_pressed("."):
                    print(".")
                    self.component_index += 1
                    self.key_flag = True
                elif keyboard.is_pressed(","):
                    print(",")
                    self.component_index -= 1
                    self.key_flag = True

                elif keyboard.is_pressed("/"):
                    print("/")
                    current_ref = self.bom_list[self.component_index]['RefDes'][0]
                    while self.bom_list[self.component_index]['RefDes'][0] == current_ref:
                        self.component_index += 1
                        if self.component_index >= len(self.bom_list):
                            self.component_index = len(self.bom_list) - 1
                            break
                    self.key_flag = True

                elif keyboard.is_pressed("m"):
                    print("m")
                    current_ref = self.bom_list[self.component_index]['RefDes'][0]
                    while self.bom_list[self.component_index]['RefDes'][0] == current_ref:
                        self.component_index -= 1
                        if self.component_index < 0:
                            self.component_index = 0
                            break
                    self.key_flag = True
                elif keyboard.is_pressed("q"):
                    print("q")
                    self.quit = True
                    return

                if self.key_flag:
                    if self.component_index < 0:
                        self.component_index = 0
                    if self.component_index >= len(self.bom_list):
                        self.component_index = len(self.bom_list) - 1
                    self.do_the_search()

    def do_the_search(self):
        self.key_flag = False

        self.close_window()

        all_windows = gw.getAllWindows()
        window_found = False
        while window_found is False:
            for window in all_windows:
                if "PCB Editor" in window.title or "PcbNew" in window.title:
                    window.activate()
                    window_found = True

        window_found = False
        while window_found is False:

            sleep(.010)
            pyautogui.hotkey('ctrl', 'f')
            sleep(.010)

            all_windows = gw.getAllWindows()
            for window in all_windows:
                if "Find" in window.title:
                    window.activate()
                    window_found = True


        pyautogui.typewrite(str(self.bom_list[self.component_index]["RefDes"]))
        pyautogui.press("enter")
        pyautogui.hotkey("shift", "tab")
        pyautogui.press("enter")
        self.window_flag = True


wind = refDesSearch()

t1 = Thread(target=wind.show_window)
t2 = Thread(target=wind.relocate_window)

t1.start()
t2.start()

wind.keyboard_input()

wind.close_window()


